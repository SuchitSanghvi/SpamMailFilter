from openpyxl import load_workbook
from mailparser import MailParser
import Tokenizer
import math


wb = load_workbook(filename='spam.xlsx', read_only=True)
ws = wb['Sheet1']

allcolumns = ws.columns

scol1v = allcolumns[0]
scol2p = allcolumns[1]

spam = {}

for i in list(range(len(scol1v))):
    spam[scol1v[i].value] = scol2p[i].value



wb = load_workbook(filename='ham.xlsx', read_only=True)
ws = wb['Sheet1']

allcolumns = ws.columns

hcol1v = allcolumns[0]
hcol2p = allcolumns[1]

ham = {}

for i in list(range(len(hcol1v))):
    ham[hcol1v[i].value] = hcol2p[i].value



parser = MailParser()
parser.parse_from_file("TRAINING/TRAIN_00549.eml")

body = parser.body

tokens = Tokenizer.freqdata(body)


PR_wi_S = 1

#for spam
for keys in tokens:
    if keys in spam.keys():
        if (float(spam.get(keys)) > 0):
            PR_wi_S = PR_wi_S * (float(spam.get(keys))*tokens.get(keys))

PR_wi_H = 1

# for ham
for keys in tokens:
    if keys in ham.keys():
        if(float(ham.get(keys))>0):
            PR_wi_H = PR_wi_H * (float(ham.get(keys))*tokens.get(keys))

print("P(E|S) = "+str(PR_wi_S))
print("P(E|H) = "+str(PR_wi_H))

num = PR_wi_S*(16/(16+25))
den = PR_wi_H*(25/(16+25))

print("P(E|S)*P(S) = "+str(num))
print("P(E|H)*P(H) = "+str(den))

r = num/den

print("P(S|E)/P(H|E) = P(E|S)*P(S)/P(E|H)*P(H) = "+str(r))

print("log(P(S|E)/P(H|E)) = "+str(math.log(r)))

if(math.log(r)>0):
    print("->>mail is spam")
else:
    print("->>mail is ham")